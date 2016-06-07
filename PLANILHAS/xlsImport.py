def xlsImport(tFile):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tFile)
        return wb
    except:
        print("Error loading xlsx file")
        raise

def getVars(ws):
    try:
        variables = {}
        print(" [#] Importing variables from first row of "+ws.title)
        count = {}
        for cell in ws.rows[0]:
            if cell.value in list(variables.values()):
                variables[cell.column] = cell.value+" "+str(count[cell.value])
                count[cell.value] += 1
            elif cell.value:
                variables[cell.column] = cell.value
                count[cell.value] = 1
    except Exception as e:
        print(" [!] Error selecting variables")
        raise
    else:
        return variables


def xlsDict(ws, variables):
    rows = []
    for row in ws.rows[1:]:
        cells = {}
        for cell in row:
            if cell.value and (cell.column in variables):
                cells[variables[cell.column]] = cell.value
        rows.append(cells)
    return rows


def Main(inp):
    #import argparse
    #parser = argparse.ArgumentParser()
    #parser.add_argument('-p', type=int, default=4, dest='protocol')
    #args = parser.parse_args()
    
    excel = xlsImport(str(inp))
    impt = {}
    varNames = getVars(excel[excel.sheetnames[0]])
    #varNames['T'] = "LEM"
    for (column, value) in varNames.items():
        if ("Ordem" in value) or ("GAB" in value) or ("HAB" in value) or (value == 'SEQ'):
            pass
        else:
            varNames[column] = False
    for sheet in excel:
        impt[sheet.title] = xlsDict(sheet, varNames)
    
    
    
    import pickle
    with open("savefile.txt", 'wb') as svFile:
        pickle.dump(impt, svFile, 4)

    return impt

if __name__=='__main__':
    Main('ITENS_ENEM_'+input()+'.xlsx')

if __name__=="__main__":
    Main('ITENS_ENEM_'+input()+'.xlsx')
